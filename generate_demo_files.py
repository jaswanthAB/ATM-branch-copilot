"""
Generates two extreme demo Excel files for interview demonstration:
  demo_healthy.xlsx  — well-performing network, minimal flags
  demo_crisis.xlsx   — struggling network, widespread issues

Run with: python generate_demo_files.py
"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

np.random.seed(99)

MONTHS     = pd.date_range("2023-01-01", periods=18, freq="MS")
REGIONS    = ["Northeast", "Southeast", "Midwest", "Southwest", "West"]
BRANCH_TYPES = ["Full-Service", "ATM-Only", "In-Store"]
BRANCH_TYPE_WEIGHTS = [0.55, 0.25, 0.20]
STATES = {
    "Northeast": ["NY","MA","PA","CT","NJ"],
    "Southeast": ["FL","GA","NC","VA","SC"],
    "Midwest":   ["IL","OH","MI","IN","WI"],
    "Southwest": ["TX","AZ","NM","NV","CO"],
    "West":      ["CA","WA","OR","UT","ID"],
}

def seasonal(month):
    base = [0.87,0.88,0.93,0.97,1.00,1.03,1.05,1.04,1.00,0.98,0.97,1.15]
    return base[month-1]


def build_records(mode="healthy"):
    """mode: 'healthy' or 'crisis'"""
    records = []
    bid = 2001 if mode == "healthy" else 3001
    branches_per_region = 15  # 75 total

    for region in REGIONS:
        for i in range(branches_per_region):
            btype  = np.random.choice(BRANCH_TYPES, p=BRANCH_TYPE_WEIGHTS)
            state  = np.random.choice(STATES[region])
            branch = f"BR-{bid}"
            bid   += 1

            if mode == "healthy":
                base_txn    = np.random.uniform(10000, 16000)
                base_uptime = np.random.uniform(0.968, 0.999)
                base_csat   = np.random.uniform(78, 92)
                decline     = 0.0
                revenue_mult = np.random.uniform(1.05, 1.25)
            else:  # crisis
                base_txn    = np.random.uniform(2500, 6000)
                base_uptime = np.random.uniform(0.72, 0.88)
                base_csat   = np.random.uniform(42, 62)
                decline     = np.random.uniform(0.02, 0.04)
                revenue_mult = np.random.uniform(0.55, 0.78)

            if btype == "ATM-Only":
                base_txn *= 0.45
            elif btype == "In-Store":
                base_txn *= 0.70

            base_cost = {"Full-Service":95000,"ATM-Only":12000,"In-Store":35000}[btype]
            base_cost *= np.random.uniform(0.85,1.15)

            for t, month_dt in enumerate(MONTHS):
                s = seasonal(month_dt.month)
                trend   = (1 - decline) ** t
                txn     = int(base_txn * s * trend * np.random.uniform(0.93,1.07))
                uptime  = round(min(1.0, max(0.5, base_uptime + np.random.normal(0,0.008))), 4)
                csat    = round(min(100, max(20, base_csat - (0.3*t if mode=="crisis" else 0) + np.random.normal(0,2))), 1)
                n_atms  = np.random.randint(1,4)
                incidents = 0 if uptime > 0.97 else (np.random.randint(3,8) if mode=="crisis" else np.random.randint(0,2))
                cash    = round(np.random.uniform(0.25,0.95),2)
                foot    = int(txn * np.random.uniform(0.08,0.14)) if btype != "ATM-Only" else 0
                wait    = round(np.random.uniform(2,4) if mode=="healthy" else np.random.uniform(9,18), 1) if btype != "ATM-Only" else 0

                base_rev = {"Full-Service":np.random.uniform(120000,220000),
                            "In-Store":np.random.uniform(40000,80000),
                            "ATM-Only":np.random.uniform(8000,20000)}[btype]
                revenue = round(base_rev * s * trend * revenue_mult * np.random.uniform(0.92,1.08), 2)
                cost    = round(base_cost * s * np.random.uniform(0.97,1.03), 2)
                roi     = round((revenue - cost)/cost*100, 2)

                records.append({
                    "Branch_ID":             branch,
                    "Region":                region,
                    "State":                 state,
                    "Branch_Type":           btype,
                    "Month":                 month_dt.strftime("%Y-%m-%d"),
                    "Transaction_Volume":    txn,
                    "Foot_Traffic":          foot,
                    "Avg_Wait_Time_Min":     wait,
                    "Num_ATMs":              n_atms,
                    "ATM_Uptime_Pct":        uptime,
                    "ATM_Downtime_Incidents":incidents,
                    "ATM_Txn_Per_Day":       round(txn/30/n_atms, 1),
                    "Cash_Level_Pct":        cash,
                    "CSAT_Score":            csat,
                    "Monthly_Revenue_USD":   revenue,
                    "Monthly_Cost_USD":      cost,
                    "ROI_Pct":               roi,
                })

    return pd.DataFrame(records)


def flag_df(df):
    df = df.copy().sort_values(["Branch_ID","Month"])
    df["Month"] = pd.to_datetime(df["Month"])

    df["Txn_3M_Rolling_Avg"] = df.groupby("Branch_ID")["Transaction_Volume"].transform(
        lambda x: x.rolling(3, min_periods=1).mean().round(0))
    df["Txn_MoM_Change_Pct"] = df.groupby("Branch_ID")["Transaction_Volume"].transform(
        lambda x: x.pct_change()*100).round(2)
    df["ATM_Uptime_3M_Avg"]  = df.groupby("Branch_ID")["ATM_Uptime_Pct"].transform(
        lambda x: x.rolling(3, min_periods=1).mean().round(4))
    df["CSAT_3M_Change"]     = df.groupby("Branch_ID")["CSAT_Score"].transform(
        lambda x: x - x.shift(3)).round(1)
    df["Revenue_Cost_Gap"]   = (df["Monthly_Revenue_USD"] - df["Monthly_Cost_USD"]).round(2)

    df["_sd"] = (df["Txn_3M_Rolling_Avg"]>0) & \
                ((df["Transaction_Volume"]-df["Txn_3M_Rolling_Avg"])/df["Txn_3M_Rolling_Avg"] < -0.15)
    df["_pd"] = df.groupby("Branch_ID")["_sd"].shift(1).fillna(False)

    df["Flag_Txn_Decline"]  = df["_sd"] & df["_pd"]
    df["Flag_ATM_Uptime"]   = df["ATM_Uptime_Pct"] < 0.90
    df["Flag_CSAT_Low"]     = df["CSAT_Score"] < 65
    df["Flag_CSAT_Drop"]    = df["CSAT_3M_Change"] < -12
    df["Flag_Negative_ROI"] = df["ROI_Pct"] < -20
    df["Is_Flagged"]        = (df["Flag_Txn_Decline"] | df["Flag_ATM_Uptime"] |
                                df["Flag_CSAT_Low"] | df["Flag_CSAT_Drop"] | df["Flag_Negative_ROI"])

    def reasons(r):
        out = []
        if r["Flag_Txn_Decline"]:  out.append("Transaction decline >15%")
        if r["Flag_ATM_Uptime"]:   out.append("ATM uptime <90%")
        if r["Flag_CSAT_Low"]:     out.append("CSAT below 65")
        if r["Flag_CSAT_Drop"]:    out.append("CSAT dropped 12+ pts")
        if r["Flag_Negative_ROI"]: out.append("Negative ROI")
        return "; ".join(out)

    df["Flag_Reasons"] = df.apply(reasons, axis=1)
    return df.drop(columns=["_sd","_pd"])


def build_sheets(df):
    cols = ["Branch_ID","Region","State","Branch_Type","Month",
            "Transaction_Volume","Txn_MoM_Change_Pct","Txn_3M_Rolling_Avg",
            "Foot_Traffic","Avg_Wait_Time_Min","Num_ATMs","ATM_Uptime_Pct",
            "ATM_Uptime_3M_Avg","ATM_Downtime_Incidents","ATM_Txn_Per_Day",
            "Cash_Level_Pct","CSAT_Score","CSAT_3M_Change","Monthly_Revenue_USD",
            "Monthly_Cost_USD","ROI_Pct","Revenue_Cost_Gap","Is_Flagged","Flag_Reasons"]

    monthly = df[cols].copy()
    monthly["Month"] = monthly["Month"].dt.strftime("%Y-%m-%d")

    latest  = df.sort_values("Month").groupby("Branch_ID").last().reset_index()
    bs_cols = ["Branch_ID","Region","State","Branch_Type","Transaction_Volume",
               "Txn_MoM_Change_Pct","ATM_Uptime_Pct","CSAT_Score","ROI_Pct",
               "Monthly_Revenue_USD","Monthly_Cost_USD","Is_Flagged","Flag_Reasons"]
    branch_summary = latest[bs_cols].copy()
    fc = df[df["Is_Flagged"]].groupby("Branch_ID").size().reset_index(name="Total_Flag_Months")
    branch_summary = branch_summary.merge(fc, on="Branch_ID", how="left")
    branch_summary["Total_Flag_Months"] = branch_summary["Total_Flag_Months"].fillna(0).astype(int)
    branch_summary = branch_summary.sort_values("Total_Flag_Months", ascending=False)

    flagged = df[df["Is_Flagged"]][cols].copy()
    flagged["Month"] = flagged["Month"].dt.strftime("%Y-%m-%d")

    region_summary = df.groupby(["Region", df["Month"].dt.to_period("M").astype(str)]).agg(
        Total_Transactions=("Transaction_Volume","sum"),
        Avg_ATM_Uptime_Pct=("ATM_Uptime_Pct","mean"),
        Avg_CSAT_Score=("CSAT_Score","mean"),
        Total_Revenue_USD=("Monthly_Revenue_USD","sum"),
        Total_Cost_USD=("Monthly_Cost_USD","sum"),
        Branches_Flagged=("Is_Flagged","sum"),
        Total_Branches=("Branch_ID","nunique"),
    ).reset_index()
    region_summary.columns.name = None
    region_summary = region_summary.rename(columns={"Month":"Period"})
    region_summary["Avg_ATM_Uptime_Pct"] = region_summary["Avg_ATM_Uptime_Pct"].round(4)
    region_summary["Avg_CSAT_Score"]     = region_summary["Avg_CSAT_Score"].round(1)
    region_summary["ROI_Pct"] = (
        (region_summary["Total_Revenue_USD"]-region_summary["Total_Cost_USD"])
        /region_summary["Total_Cost_USD"]*100).round(2)

    return monthly, branch_summary, flagged, region_summary


def write_excel(monthly, branch_summary, flagged, region_summary, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        monthly.to_excel(writer,        sheet_name="Monthly_Data",    index=False)
        branch_summary.to_excel(writer, sheet_name="Branch_Summary",  index=False)
        flagged.to_excel(writer,        sheet_name="Flagged_Branches", index=False)
        region_summary.to_excel(writer, sheet_name="Region_Summary",   index=False)

    wb = load_workbook(path)
    hdr = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = hdr; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for col_idx, col in enumerate(ws.columns, 1):
            w = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(w+4, 36)
        ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    for mode, fname in [("healthy","demo_healthy.xlsx"), ("crisis","demo_crisis.xlsx")]:
        print(f"Building {mode} dataset...")
        raw = build_records(mode)
        df  = flag_df(raw)
        mo, bs, fl, rs = build_sheets(df)
        write_excel(mo, bs, fl, rs, fname)
        flagged_n = bs["Is_Flagged"].sum()
        print(f"  Branches: {bs['Branch_ID'].nunique()} | Flagged: {flagged_n} | Avg CSAT: {bs['CSAT_Score'].mean():.1f} | Avg ATM Uptime: {bs['ATM_Uptime_Pct'].mean()*100:.1f}%")
        print(f"  Saved → {fname}\n")

    print("Done. Upload demo_healthy.xlsx or demo_crisis.xlsx into Branch Copilot to compare.")
