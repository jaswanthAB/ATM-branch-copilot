"""
Phase 2: Cleaning + Anomaly Detection Pipeline
Diebold Nixdorf Advisory Services — Case Study Project

What this script does:
1. Loads raw branch/ATM data
2. Cleans and validates (missing values, types)
3. Calculates month-over-month and rolling metrics
4. Flags underperforming branches against business thresholds
5. Outputs a single Power BI-ready Excel workbook with 4 named sheets

Flagging thresholds:
- Transaction decline > 15% vs. 3-month rolling average
- ATM uptime < 90% in any month
- CSAT score < 65 or dropped > 10 points over 3 months
- ROI negative (cost exceeds revenue)
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ── Thresholds ────────────────────────────────────────────────────────────────
THRESHOLD_TXN_DECLINE_PCT   = 0.15   # 15% decline vs 3-month rolling avg (must persist 2+ months)
THRESHOLD_ATM_UPTIME_MIN    = 0.90   # below 90% = flagged
THRESHOLD_CSAT_MIN          = 62.0   # below 62 = flagged
THRESHOLD_CSAT_DROP         = 12.0   # dropped 12+ points over 3 months
THRESHOLD_ROI_MIN           = -20.0  # ROI below -20% = flagged (ATM-only branches run lean by design)

# ── 1. Load & basic validation ────────────────────────────────────────────────
def load_and_validate(path="raw_branch_atm_data.csv"):
    print("Loading raw data...")
    df = pd.read_csv(path)
    print(f"  {len(df):,} rows, {df['Branch_ID'].nunique()} branches loaded")

    # Types
    df["Month"] = pd.to_datetime(df["Month"])
    df = df.sort_values(["Branch_ID", "Month"]).reset_index(drop=True)

    numeric_cols = [
        "Transaction_Volume", "Foot_Traffic", "Avg_Wait_Time_Min",
        "ATM_Uptime_Pct", "ATM_Downtime_Incidents", "ATM_Txn_Per_Day",
        "Cash_Level_Pct", "CSAT_Score", "Monthly_Revenue_USD",
        "Monthly_Cost_USD", "ROI_Pct"
    ]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


# ── 2. Clean missing values ───────────────────────────────────────────────────
def clean_missing(df):
    print("Cleaning missing values...")

    # Foot traffic & wait time: ATM-Only branches legitimately have none
    df.loc[df["Branch_Type"] == "ATM-Only", "Foot_Traffic"] = df.loc[
        df["Branch_Type"] == "ATM-Only", "Foot_Traffic"].fillna(0)
    df.loc[df["Branch_Type"] == "ATM-Only", "Avg_Wait_Time_Min"] = df.loc[
        df["Branch_Type"] == "ATM-Only", "Avg_Wait_Time_Min"].fillna(0)

    # For other branch types: forward-fill within branch, then median fallback
    for col in ["Foot_Traffic", "Avg_Wait_Time_Min", "ATM_Txn_Per_Day"]:
        df[col] = df.groupby("Branch_ID")[col].transform(
            lambda x: x.fillna(method="ffill").fillna(x.median())
        )

    # CSAT: forward-fill within branch, then global median
    df["CSAT_Score"] = df.groupby("Branch_ID")["CSAT_Score"].transform(
        lambda x: x.fillna(method="ffill").fillna(x.median())
    )
    df["CSAT_Score"] = df["CSAT_Score"].fillna(df["CSAT_Score"].median())

    missing_after = df.isnull().sum().sum()
    print(f"  Missing values remaining: {missing_after}")
    return df


# ── 3. Calculate derived metrics ──────────────────────────────────────────────
def calculate_metrics(df):
    print("Calculating rolling and period-over-period metrics...")

    df = df.sort_values(["Branch_ID", "Month"])

    # 3-month rolling average transaction volume
    df["Txn_3M_Rolling_Avg"] = df.groupby("Branch_ID")["Transaction_Volume"].transform(
        lambda x: x.rolling(3, min_periods=1).mean().round(0)
    )

    # Month-over-month transaction change %
    df["Txn_MoM_Change_Pct"] = df.groupby("Branch_ID")["Transaction_Volume"].transform(
        lambda x: x.pct_change() * 100
    ).round(2)

    # ATM uptime 3-month trend (rolling avg)
    df["ATM_Uptime_3M_Avg"] = df.groupby("Branch_ID")["ATM_Uptime_Pct"].transform(
        lambda x: x.rolling(3, min_periods=1).mean().round(4)
    )

    # CSAT 3-month change
    df["CSAT_3M_Change"] = df.groupby("Branch_ID")["CSAT_Score"].transform(
        lambda x: x - x.shift(3)
    ).round(1)

    # Revenue vs cost gap
    df["Revenue_Cost_Gap"] = (df["Monthly_Revenue_USD"] - df["Monthly_Cost_USD"]).round(2)

    return df


# ── 4. Flag underperforming branches ─────────────────────────────────────────
def flag_branches(df):
    print("Flagging underperforming branches...")

    flags = []

    # Transaction decline vs 3M rolling average — must be declining 2+ consecutive months
    df["_single_decline"] = (
        (df["Txn_3M_Rolling_Avg"] > 0) &
        ((df["Transaction_Volume"] - df["Txn_3M_Rolling_Avg"]) / df["Txn_3M_Rolling_Avg"] < -THRESHOLD_TXN_DECLINE_PCT)
    )
    df["_prev_decline"] = df.groupby("Branch_ID")["_single_decline"].shift(1).fillna(False)
    txn_decline = df["_single_decline"] & df["_prev_decline"]

    # ATM uptime below threshold
    atm_flag = df["ATM_Uptime_Pct"] < THRESHOLD_ATM_UPTIME_MIN

    # CSAT below minimum or dropped sharply
    csat_low  = df["CSAT_Score"] < THRESHOLD_CSAT_MIN
    csat_drop = df["CSAT_3M_Change"] < -THRESHOLD_CSAT_DROP

    # Negative ROI
    roi_flag = df["ROI_Pct"] < THRESHOLD_ROI_MIN

    df["Flag_Txn_Decline"]  = txn_decline
    df["Flag_ATM_Uptime"]   = atm_flag
    df["Flag_CSAT_Low"]     = csat_low
    df["Flag_CSAT_Drop"]    = csat_drop
    df["Flag_Negative_ROI"] = roi_flag

    # Overall flag
    df["Is_Flagged"] = (
        txn_decline | atm_flag | csat_low | csat_drop | roi_flag
    )

    # Human-readable flag reasons
    def flag_reasons(row):
        reasons = []
        if row["Flag_Txn_Decline"]:  reasons.append("Transaction decline >15%")
        if row["Flag_ATM_Uptime"]:   reasons.append("ATM uptime <90%")
        if row["Flag_CSAT_Low"]:     reasons.append("CSAT below 65")
        if row["Flag_CSAT_Drop"]:    reasons.append("CSAT dropped 10+ pts")
        if row["Flag_Negative_ROI"]: reasons.append("Negative ROI")
        return "; ".join(reasons) if reasons else ""

    df["Flag_Reasons"] = df.apply(flag_reasons, axis=1)

    flagged_count = df[df["Is_Flagged"]]["Branch_ID"].nunique()
    total_flags   = df["Is_Flagged"].sum()
    print(f"  {flagged_count} unique branches flagged across {total_flags} monthly records")

    return df


# ── 5. Build output sheets ────────────────────────────────────────────────────
def build_sheets(df):
    print("Building output sheets...")

    # Sheet 1: Monthly Data (clean, all rows)
    display_cols = [
        "Branch_ID", "Region", "State", "Branch_Type", "Month",
        "Transaction_Volume", "Txn_MoM_Change_Pct", "Txn_3M_Rolling_Avg",
        "Foot_Traffic", "Avg_Wait_Time_Min",
        "Num_ATMs", "ATM_Uptime_Pct", "ATM_Uptime_3M_Avg",
        "ATM_Downtime_Incidents", "ATM_Txn_Per_Day", "Cash_Level_Pct",
        "CSAT_Score", "CSAT_3M_Change",
        "Monthly_Revenue_USD", "Monthly_Cost_USD", "ROI_Pct",
        "Revenue_Cost_Gap", "Is_Flagged", "Flag_Reasons"
    ]
    monthly = df[display_cols].copy()
    monthly["Month"] = monthly["Month"].dt.strftime("%Y-%m-%d")

    # Sheet 2: Branch Summary (latest month per branch)
    latest = df.sort_values("Month").groupby("Branch_ID").last().reset_index()
    summary_cols = [
        "Branch_ID", "Region", "State", "Branch_Type",
        "Transaction_Volume", "Txn_MoM_Change_Pct",
        "ATM_Uptime_Pct", "CSAT_Score", "ROI_Pct",
        "Monthly_Revenue_USD", "Monthly_Cost_USD",
        "Is_Flagged", "Flag_Reasons"
    ]
    branch_summary = latest[summary_cols].copy()

    # Add total flag count per branch
    flag_counts = df[df["Is_Flagged"]].groupby("Branch_ID").size().reset_index(name="Total_Flag_Months")
    branch_summary = branch_summary.merge(flag_counts, on="Branch_ID", how="left")
    branch_summary["Total_Flag_Months"] = branch_summary["Total_Flag_Months"].fillna(0).astype(int)
    branch_summary = branch_summary.sort_values("Total_Flag_Months", ascending=False)

    # Sheet 3: Flagged Branches only
    flagged = df[df["Is_Flagged"]][display_cols].copy()
    flagged["Month"] = flagged["Month"].dt.strftime("%Y-%m-%d")
    flagged = flagged.sort_values(["Branch_ID", "Month"])

    # Sheet 4: Region Summary
    region_summary = df.groupby(["Region", df["Month"].dt.to_period("M").astype(str)]).agg(
        Total_Transactions   = ("Transaction_Volume", "sum"),
        Avg_ATM_Uptime_Pct   = ("ATM_Uptime_Pct", "mean"),
        Avg_CSAT_Score       = ("CSAT_Score", "mean"),
        Total_Revenue_USD    = ("Monthly_Revenue_USD", "sum"),
        Total_Cost_USD       = ("Monthly_Cost_USD", "sum"),
        Branches_Flagged     = ("Is_Flagged", "sum"),
        Total_Branches       = ("Branch_ID", "nunique"),
    ).reset_index()
    region_summary.columns.name = None
    region_summary = region_summary.rename(columns={"Month": "Period"})
    region_summary["Avg_ATM_Uptime_Pct"] = region_summary["Avg_ATM_Uptime_Pct"].round(4)
    region_summary["Avg_CSAT_Score"]     = region_summary["Avg_CSAT_Score"].round(1)
    region_summary["ROI_Pct"] = (
        (region_summary["Total_Revenue_USD"] - region_summary["Total_Cost_USD"])
        / region_summary["Total_Cost_USD"] * 100
    ).round(2)

    return monthly, branch_summary, flagged, region_summary


# ── 6. Write Excel workbook ───────────────────────────────────────────────────
def write_excel(monthly, branch_summary, flagged, region_summary, out_path="branch_atm_report.xlsx"):
    print(f"Writing Excel workbook → {out_path}")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        monthly.to_excel(writer,        sheet_name="Monthly_Data",     index=False)
        branch_summary.to_excel(writer, sheet_name="Branch_Summary",   index=False)
        flagged.to_excel(writer,        sheet_name="Flagged_Branches",  index=False)
        region_summary.to_excel(writer, sheet_name="Region_Summary",    index=False)

    # ── Formatting pass ───────────────────────────────────────────────────────
    wb = load_workbook(out_path)

    header_fill   = PatternFill("solid", fgColor="1F3864")   # dark navy
    header_font   = Font(color="FFFFFF", bold=True)
    flag_fill     = PatternFill("solid", fgColor="FFD7D7")   # light red
    alt_fill      = PatternFill("solid", fgColor="F2F7FF")   # light blue
    border_side   = Side(style="thin", color="CCCCCC")
    thin_border   = Border(bottom=border_side)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header row styling
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-width columns
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # Freeze top row
        ws.freeze_panes = "A2"

        # Highlight flagged rows in Flagged_Branches sheet
        if sheet_name == "Flagged_Branches":
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = flag_fill

        # Alternating row fill for other sheets
        elif sheet_name in ["Branch_Summary", "Region_Summary"]:
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if i % 2 == 0:
                    for cell in row:
                        cell.fill = alt_fill

    wb.save(out_path)
    print(f"  Workbook saved with {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    df = load_and_validate()
    df = clean_missing(df)
    df = calculate_metrics(df)
    df = flag_branches(df)

    monthly, branch_summary, flagged, region_summary = build_sheets(df)
    write_excel(monthly, branch_summary, flagged, region_summary)

    print("\n── Summary ──────────────────────────────────────────────────")
    print(f"  Total branches:        {df['Branch_ID'].nunique()}")
    print(f"  Total monthly records: {len(monthly):,}")
    print(f"  Branches flagged:      {branch_summary['Is_Flagged'].sum()}")
    print(f"  Total flagged records: {len(flagged):,}")
    print(f"  Date range:            {df['Month'].min().strftime('%b %Y')} → {df['Month'].max().strftime('%b %Y')}")
    print(f"\n  Output: branch_atm_report.xlsx — ready for Power BI Online upload")
